import re
from pathlib import Path
from typing import List, Union

from openpyxl import load_workbook

from pecha_org_categorizer.config import PECHA_CATEGORIES_FILE
from pecha_org_categorizer.enums import TextType


class CategoryExtractor:
    def __init__(self, input_file: Path = PECHA_CATEGORIES_FILE):
        self.input_file = input_file
        self.bo_formatted_categories, self.en_formatted_categories = self.process_file()

    @staticmethod
    def read_xlsx_file(file_path: Path):
        """
        Read the xlsx file and return its contents as a list of rows.
        """
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        rows_data = []
        for row in worksheet.iter_rows(values_only=True):
            rows_data.append(row)
        return rows_data

    def extract_categories(self):
        """
        Process the xlsx file and extract hierarchical categories from its contents.
        """
        self.rows_data = self.read_xlsx_file(self.input_file)
        extracted_categories = []
        current_category: List[Union[str, None]] = []

        en_extracted_categories = []
        current_en_category: List[Union[str, None]] = []

        for row in self.rows_data:
            # Find the first non-None value and its index
            flag = False
            for col_index, cell_value in enumerate(row):
                if cell_value is not None:
                    flag = True
                    break
            if not flag:
                continue

            current_category_len = len(current_category)
            cell_value = cell_value.strip()

            en_cell_value = row[col_index + 1].strip()

            # Update or extend the current category hierarchy
            if current_category_len == col_index:
                current_category.append(cell_value)
                current_en_category.append(en_cell_value)
            else:
                current_category[col_index] = cell_value
                current_en_category[col_index] = en_cell_value

            # Reset trailing elements to None
            current_category[col_index + 1 :] = [None] * (  # noqa
                current_category_len - col_index - 1
            )
            current_en_category[col_index + 1 :] = [None] * (  # noqa
                current_category_len - col_index - 1
            )

            # Add non-empty elements to the result
            active_category = [
                category for category in current_category if category is not None
            ]
            extracted_categories.append(active_category)

            active_en_category = [
                category for category in current_en_category if category is not None
            ]
            en_extracted_categories.append(active_en_category)

        self.bo_extracted_categories = extracted_categories
        self.en_extracted_categories = en_extracted_categories

    def process_file(self):
        """
        Extract and format categories from the provided xlsx file.
        """
        self.extract_categories()
        bo_formatted_categories, en_formatted_categories = [], []
        for bo_category_hierarchy, en_category_hierarchy in zip(
            self.bo_extracted_categories, self.en_extracted_categories
        ):
            bo_formatted_category = format_categories(bo_category_hierarchy, "bo")
            en_formatted_category = format_categories(en_category_hierarchy, "en")

            bo_formatted_categories.append(bo_formatted_category)
            en_formatted_categories.append(en_formatted_category)

        return bo_formatted_categories, en_formatted_categories

    def get_category_hierarchy(
        self,
        category_name: str,
        pecha_metadata: dict,
        lang: str,
        text_type: TextType = TextType.NONE,
    ):
        """
        Get the category hierarchy for a given category name.
        """
        assert "title" in pecha_metadata
        assert "heDesc" in pecha_metadata or "enDesc" in pecha_metadata
        assert "heShortDesc" in pecha_metadata or "enShortDesc" in pecha_metadata

        if lang == "bo":
            formatted_categories = self.bo_formatted_categories
        elif lang == "en":
            formatted_categories = self.en_formatted_categories
        else:
            raise ValueError(f"Unsupported language: {lang}")

        matched_category_hierarchy = None
        for category_hierarchy in formatted_categories:
            for category in category_hierarchy:
                if category["name"] == category_name:
                    matched_category_hierarchy = category_hierarchy
                    break
            if matched_category_hierarchy:
                break

        if matched_category_hierarchy is None:
            raise ValueError(f"Category not found for {category_name}")

        if lang == "bo":
            if text_type == TextType.ROOT:
                matched_category_hierarchy.append(
                    {"name": "རྩ་བ།", "heDesc": "", "heShortDesc": ""}
                )
            elif text_type == TextType.COMMENTARY:
                matched_category_hierarchy.append(
                    {"name": "འགྲེལ་པ།", "heDesc": "", "heShortDesc": ""}
                )
            matched_category_hierarchy.append(
                {
                    "name": pecha_metadata["title"],
                    "heDesc": pecha_metadata["heDesc"],
                    "heShortDesc": pecha_metadata["heShortDesc"],
                }
            )
        else:
            if text_type == TextType.ROOT:
                matched_category_hierarchy.append(
                    {"name": "Root text", "enDesc": "", "enShortDesc": ""}
                )
            elif text_type == TextType.COMMENTARY:
                matched_category_hierarchy.append(
                    {"name": "Commentaries", "enDesc": "", "enShortDesc": ""}
                )
            matched_category_hierarchy.append(
                {
                    "name": pecha_metadata["title"],
                    "enDesc": pecha_metadata["enDesc"],
                    "enShortDesc": pecha_metadata["enShortDesc"],
                }
            )
        return matched_category_hierarchy


def extract_text_details(text: str):
    """
    Extract the main text and any descriptions (in parentheses) from a given string.
    """
    pattern = r"^(.*?)\s*(?:\((.*?)\))?(?:\((.*?)\))?$"
    match = re.search(pattern, text)

    if match:
        name = match.group(1).strip() if match.group(1) else ""
        description = match.group(2).strip() if match.group(2) else ""
        short_description = match.group(3).strip() if match.group(3) else ""
        return name, description, short_description
    else:
        return None, None, None


def format_categories(category_hierarchy: List[str], lang: str):
    """
    Format each category hierarchy into a structured format with main text and descriptions.
    """
    formatted_hierarchy = []
    for category in category_hierarchy:
        name, description, short_description = extract_text_details(category)
        if lang == "bo":
            category_data = {
                "name": name,
                "heDesc": description,
                "heShortDesc": short_description,
            }
        else:
            category_data = {
                "name": name,
                "enDesc": description,
                "enShortDesc": short_description,
            }

        formatted_hierarchy.append(category_data)

    return formatted_hierarchy
